"""Runs the simulations."""
from pathlib import Path
import time

import numpy as np
from openpyxl import Workbook

from scenarios import scenarios, cycle_length, Stats


def simulate_period(weeks, scenario):
    """Runs a simulation over the defined period.

        Attributes:
            weeks (float): the number of weeks in each simulation
            case_details (dict): a dicationary of details for this case
    """
    # The employee capacity for shifts (or how many shifts the staff can
    # cover in this scenario)
    shift_capacity = scenario.fte.actual.total * 5

    # Dictionary to contain all event outcomes
    event_results = {}

    for event in scenario.events:
        event_results[event.name] = {
            'outcome_0': 0,
            'outcome_2': 0,
            'outcome_4': 0,
            'outcome_12': 0,
            'outcome_total': 0,
        }

    # Dictionary to hold all results for shift evaluations
    uncovered_shifts = {}
    excess_shifts = 0

    for shift in scenario.shifts:
        uncovered_shifts[shift.name] = 0

    # Variable to track the number of shift changes
    total_shift_changes = {
        'change_0': 0,
        'change_2': 0,
        'change_4': 0,
        'change_12': 0,
        'change_total': 0,
    }

    # Iterate through each week
    for _ in range(weeks):
        # Numpy generator for binomial evaluations
        gen = np.random.Generator(np.random.PCG64())

        # Number of shifts lost each week
        week_shift_losses = 0

        for event in scenario.events:
            # Only evaluate event if current total is less than cycle max
            if event.cycle_max and event_results[event.name]['outcome_total'] >= event.cycle_max:
                continue
            else:
                # Check if an event occurence happens for each timeframe
                week_0 = gen.binomial(shift_capacity, event.rate_0)
                week_2 = gen.binomial(shift_capacity, event.rate_2)
                week_4 = gen.binomial(shift_capacity, event.rate_4)
                week_12 = gen.binomial(shift_capacity, event.rate_12)
                event_total = week_0 + week_2 + week_4 + week_12

                week_shift_losses += (event_total * event.losses)

                week_changes_0 = week_0 * event.changes
                week_changes_2 = week_2 * event.changes
                week_changes_4 = week_4 * event.changes
                week_changes_12 = week_12 * event.changes
                week_changes_total = event_total * event.changes

                # Update the results dictionaries with this week's results
                event_results[event.name]['outcome_0'] += week_0
                event_results[event.name]['outcome_2'] += week_2
                event_results[event.name]['outcome_4'] += week_4
                event_results[event.name]['outcome_12'] += week_12
                event_results[event.name]['outcome_total'] += event_total

                total_shift_changes['change_0'] += week_changes_0
                total_shift_changes['change_2'] += week_changes_2
                total_shift_changes['change_4'] += week_changes_4
                total_shift_changes['change_12'] += week_changes_12
                total_shift_changes['change_total'] += week_changes_total

        # Evaluates how many shifts can be covered in this scenario based
        # based on desired shifts to be covered, the employee availability,
        # and the events that occurred this week.
        # The starting capacity will be the normal weekly capacity minus the
        # week event total.
        remaining_capacity = shift_capacity - week_shift_losses

        # Iterate through each shift and start assigning remaining capacity
        for shift in scenario.shifts:
            if remaining_capacity >= shift.number:
                remaining_capacity -= shift.number
            else:
                remaining_capacity = 0
                uncovered_shifts[shift.name] += (shift.number - remaining_capacity)

        # Record any remaining capacity
        if remaining_capacity > 0:
            excess_shifts += remaining_capacity

    return {
        'events': event_results,
        'uncovered_shifts': uncovered_shifts,
        'remaining_capacity': excess_shifts,
        'shift_changes': total_shift_changes,
    }

# Simulation Details
num_simulations = 1000

# Running the simulations
print('========================================================================')
print('RDRHC Monte Carlo Simulations')
print('========================================================================')
print('Created by Joshua Torrance. 2024.')
print('\nRunning simulations for each scenario.')
print(f'  - Number of Simulations per Scenario: {num_simulations}')
print(f'  - Length of Each Simulation: {cycle_length} weeks')
print('\n------------------------------------------------------------------------')
print('SCENARIOS')
print('------------------------------------------------------------------------')

# Variables to manage output workbook details
output_wb = Workbook()
output_ws = output_wb.active
num_scenarios = len(scenarios)

for scenario in scenarios:
    print(f'  - {scenario.name}')

    # Set Worksheet Title
    output_ws.title = scenario.name

    # Run simulations function and collect results in array
    scenario_results = []

    for _ in range(num_simulations):
        scenario_results.append(simulate_period(cycle_length, scenario))

    # Analyze event results
    # Add each type of event as a dictionary value
    simulation_event_results = {
        'All Events': {
            'outcome_0': [],
            'outcome_2': [],
            'outcome_4': [],
            'outcome_12': [],
            'outcome_total': [],
        },
    }

    for event in scenario.events:
        simulation_event_results[event.name] = {
            'outcome_0': [],
            'outcome_2': [],
            'outcome_4': [],
            'outcome_12': [],
            'outcome_total': [],
        }

    # Add each type of shift as a dictionary value
    simulation_coverage_results = {
        'All Shifts': [],
    }

    for shift in scenario.shifts:
        simulation_coverage_results[shift.name] = []
    
    # Holds the count for number of excess shifts
    simulation_excess_shifts = []

    # Holds the number of shift changes
    simulation_shift_changes = {
        'week_0': [],
        'week_2': [],
        'week_4': [],
        'week_12': [],
        'week_total': [],
    }

    # Iterate through each scenario results
    for cycle_results in scenario_results:
        all_events_0 = []
        all_events_2 = []
        all_events_4 = []
        all_events_12 = []
        all_events_total = []
        all_uncovered_shifts = 0

        # Iterate through the results of each event
        for name, results in cycle_results['events'].items():
            outcome_0 = results['outcome_0']
            outcome_2 = results['outcome_2']
            outcome_4 = results['outcome_4']
            outcome_12 = results['outcome_12']
            outcome_total = results['outcome_total']

            simulation_event_results[name]['outcome_0'].append(outcome_0)
            simulation_event_results[name]['outcome_2'].append(outcome_2)
            simulation_event_results[name]['outcome_4'].append(outcome_4)
            simulation_event_results[name]['outcome_12'].append(outcome_12)
            simulation_event_results[name]['outcome_total'].append(outcome_total)

            all_events_0.append(outcome_0)
            all_events_2.append(outcome_2)
            all_events_4.append(outcome_4)
            all_events_12.append(outcome_12)
            all_events_total.append(outcome_total)
        
        # Update the special "All Events" entry
        simulation_event_results['All Events']['outcome_0'] = all_events_0
        simulation_event_results['All Events']['outcome_2'] = all_events_2
        simulation_event_results['All Events']['outcome_4'] = all_events_4
        simulation_event_results['All Events']['outcome_12'] = all_events_12
        simulation_event_results['All Events']['outcome_total'] = all_events_total

        # Iterate through the results of each shift
        for name, results in cycle_results['uncovered_shifts'].items():
            simulation_coverage_results[name].append(results)
            all_uncovered_shifts += results

        # Update the special "All Shifts" entry
        simulation_coverage_results['All Shifts'].append(all_uncovered_shifts)

        # Add the excess shifts results
        simulation_excess_shifts.append(cycle_results['remaining_capacity'])

        # Add the shift changes results
        simulation_shift_changes['week_0'].append(cycle_results['shift_changes']['change_0'])
        simulation_shift_changes['week_2'].append(cycle_results['shift_changes']['change_2'])
        simulation_shift_changes['week_4'].append(cycle_results['shift_changes']['change_4'])
        simulation_shift_changes['week_12'].append(cycle_results['shift_changes']['change_12'])
        simulation_shift_changes['week_total'].append(cycle_results['shift_changes']['change_total'])

    # Iterate through the lists of simulation event results to run calculations
    simulations_stats = {
        'events': {},
        'uncovered_shifts': {},
    }

    for name, values in simulation_event_results.items():
        stats_0 = Stats(values['outcome_0'])
        stats_2 = Stats(values['outcome_2'])
        stats_4 = Stats(values['outcome_4'])
        stats_12 = Stats(values['outcome_12'])
        stats_total = Stats(values['outcome_total'])

        simulations_stats['events'][name] = {
            'stats_0': stats_0,
            'stats_2': stats_2,
            'stats_4': stats_4,
            'stats_12': stats_12,
            'stats_total': stats_total,
        }

    # Shift Coverage Stats
    for name, values in simulation_coverage_results.items():
        uncovered_shift_stats = Stats(values)

        simulations_stats['uncovered_shifts'][name] = uncovered_shift_stats
        
    # Excess shift capacity stats
    excess_shifts_stats = Stats(simulation_excess_shifts)

    # Number of shift changes for cycle
    shift_changes_stats_0 = Stats(simulation_shift_changes['week_0'])
    shift_changes_stats_2 = Stats(simulation_shift_changes['week_2'])
    shift_changes_stats_4 = Stats(simulation_shift_changes['week_4'])
    shift_changes_stats_12 = Stats(simulation_shift_changes['week_12'])
    shift_changes_stats_total = Stats(simulation_shift_changes['week_total'])

    # Write data to the active worksheet
    row_num = 1

    # Basic Monte Carlo Details
    output_ws.cell(row=row_num, column=1, value='SIMULATION DETAILS')
    row_num += 1
    
    output_ws.cell(row=row_num, column=2, value='Value')
    row_num += 1

    output_ws.cell(row=row_num, column=1, value='Number of Simulations')
    output_ws.cell(row=row_num, column=2, value=num_simulations)
    row_num += 1

    output_ws.cell(row=row_num, column=1, value='Length of Simulation Cycle (weeks)')
    output_ws.cell(row=row_num, column=2, value=cycle_length)
    row_num += 2

    # Event Details
    output_ws.cell(row=row_num, column=1, value='EVENT DETAILS')
    row_num += 1

    output_ws.cell(row=row_num, column=1, value='Event')
    output_ws.cell(row=row_num, column=2, value='Rate of Changes (changes per event occurrence)')
    output_ws.cell(row=row_num, column=3, value='Rate of Lost Shift Capacity (lost shifts per event occurrence)')
    output_ws.cell(row=row_num, column=4, value='Event Rate Occurence - Total')
    output_ws.cell(row=row_num, column=5, value='Event Rate Occurence - 0 to 2 weeks')
    output_ws.cell(row=row_num, column=6, value='Event Rate Occurence - 2 to 4 weeks')
    output_ws.cell(row=row_num, column=7, value='Event Rate Occurence - 4 to 12 weeks')
    output_ws.cell(row=row_num, column=8, value='Event Rate Occurence - 12+ weeks')
    output_ws.cell(row=row_num, column=9, value='Maximum Number of Allowed Events per Cycle')
    row_num += 1

    for event in scenario.events:
        output_ws.cell(row=row_num, column=1, value=event.name)
        output_ws.cell(row=row_num, column=2, value=event.changes)
        output_ws.cell(row=row_num, column=3, value=event.losses)
        output_ws.cell(row=row_num, column=4, value=np.round(event.rate_total, 2))
        output_ws.cell(row=row_num, column=5, value=np.round(event.rate_0, 2))
        output_ws.cell(row=row_num, column=6, value=np.round(event.rate_2, 2))
        output_ws.cell(row=row_num, column=7, value=np.round(event.rate_4, 2))
        output_ws.cell(row=row_num, column=8, value=np.round(event.rate_12, 2))
        output_ws.cell(row=row_num, column=9, value=event.cycle_max)
        row_num +=1

    row_num +=1

    # Event occurence details for the simulation
    output_ws.cell(row=row_num, column=1, value='EVENT RESULTS')
    row_num += 1

    output_ws.cell(row=row_num, column=1, value='Event')
    output_ws.cell(row=row_num, column=2, value='Total - Mean')
    output_ws.cell(row=row_num, column=3, value='Total - Lower CI')
    output_ws.cell(row=row_num, column=4, value='Total - Upper CI')
    output_ws.cell(row=row_num, column=5, value='Weeks 0 to 2 - Mean')
    output_ws.cell(row=row_num, column=6, value='Weeks 0 to 2 - Lower CI')
    output_ws.cell(row=row_num, column=7, value='Weeks 0 to 2 - Upper CI')
    output_ws.cell(row=row_num, column=8, value='Weeks 2 to 4 - Mean')
    output_ws.cell(row=row_num, column=9, value='Weeks 2 to 4 - Lower CI')
    output_ws.cell(row=row_num, column=10, value='Weeks 2 to 4 - Upper CI')
    output_ws.cell(row=row_num, column=11, value='Weeks 4 to 12 - Mean')
    output_ws.cell(row=row_num, column=12, value='Weeks 4 to 12 - Lower CI')
    output_ws.cell(row=row_num, column=13, value='Weeks 4 to 12 - Upper CI')
    output_ws.cell(row=row_num, column=14, value='Weeks 12+ - Mean')
    output_ws.cell(row=row_num, column=15, value='Weeks 12+ - Lower CI')
    output_ws.cell(row=row_num, column=16, value='Weeks 12+ - Upper CI')

    row_num += 1

    for event_name, event_stats in simulations_stats['events'].items():
        output_ws.cell(row=row_num, column=1, value=event_name)
        output_ws.cell(row=row_num, column=2, value=event_stats['stats_total'].mean)
        output_ws.cell(row=row_num, column=3, value=event_stats['stats_total'].ci_lower)
        output_ws.cell(row=row_num, column=4, value=event_stats['stats_total'].ci_upper)
        output_ws.cell(row=row_num, column=5, value=event_stats['stats_0'].mean)
        output_ws.cell(row=row_num, column=6, value=event_stats['stats_0'].ci_lower)
        output_ws.cell(row=row_num, column=7, value=event_stats['stats_0'].ci_upper)
        output_ws.cell(row=row_num, column=8, value=event_stats['stats_2'].mean)
        output_ws.cell(row=row_num, column=9, value=event_stats['stats_2'].ci_lower)
        output_ws.cell(row=row_num, column=10, value=event_stats['stats_2'].ci_upper)
        output_ws.cell(row=row_num, column=11, value=event_stats['stats_4'].mean)
        output_ws.cell(row=row_num, column=12, value=event_stats['stats_4'].ci_lower)
        output_ws.cell(row=row_num, column=13, value=event_stats['stats_4'].ci_upper)
        output_ws.cell(row=row_num, column=14, value=event_stats['stats_12'].mean)
        output_ws.cell(row=row_num, column=15, value=event_stats['stats_12'].ci_lower)
        output_ws.cell(row=row_num, column=16, value=event_stats['stats_12'].ci_upper)
        row_num += 1

    row_num += 1

    # Shift Details
    output_ws.cell(row=row_num, column=1, value='SHIFT DETAILS')
    row_num += 1

    output_ws.cell(row=row_num, column=1, value='Shift Name')
    output_ws.cell(row=row_num, column=2, value='Number of Shifts')
    output_ws.cell(row=row_num, column=3, value='Shift Priority')
    row_num += 1

    for shift in scenario.shifts:
        output_ws.cell(row=row_num, column=1, value=shift.name)
        output_ws.cell(row=row_num, column=2, value=shift.number)
        output_ws.cell(row=row_num, column=3, value=shift.priority)
        row_num += 1
        
    row_num += 1

    # Shift Results
    output_ws.cell(row=row_num, column=1, value='UNCOVERED SHIFT RESULTS')
    row_num += 1

    output_ws.cell(row=row_num, column=1, value='Shift')
    output_ws.cell(row=row_num, column=2, value='Mean Uncovered Shifts per Cycle')
    output_ws.cell(row=row_num, column=3, value='Lower CI')
    output_ws.cell(row=row_num, column=4, value='Upper CI')
    row_num += 1

    for shift_name, shift_stats in simulations_stats['uncovered_shifts'].items():
        output_ws.cell(row=row_num, column=1, value=shift_name)
        output_ws.cell(row=row_num, column=2, value=shift_stats.mean)
        output_ws.cell(row=row_num, column=3, value=shift_stats.ci_lower)
        output_ws.cell(row=row_num, column=4, value=shift_stats.ci_upper)
        row_num += 1

    row_num += 1

    # Excess Shift Results
    output_ws.cell(row=row_num, column=1, value='EXCESS SHIFT RESULTS')
    row_num += 1
    
    output_ws.cell(row=row_num, column=2, value='Mean Number of Shifts Per Cycle')
    output_ws.cell(row=row_num, column=3, value='Lower CI')
    output_ws.cell(row=row_num, column=4, value='Upper CI')
    row_num += 1
    
    output_ws.cell(row=row_num, column=1, value='Number of Excess Shifts')
    output_ws.cell(row=row_num, column=2, value=excess_shifts_stats.mean)
    output_ws.cell(row=row_num, column=3, value=excess_shifts_stats.ci_lower)
    output_ws.cell(row=row_num, column=4, value=excess_shifts_stats.ci_upper)
    row_num += 2

    # Number of Shift Changes
    output_ws.cell(row=row_num, column=1, value='NUMBER OF SHIFT CHANGES')
    row_num += 1

    output_ws.cell(row=row_num, column=2, value='Total - Mean')
    output_ws.cell(row=row_num, column=3, value='Total - Lower CI')
    output_ws.cell(row=row_num, column=4, value='Total - Upper CI')
    output_ws.cell(row=row_num, column=5, value='Weeks 0 to 2 - Mean')
    output_ws.cell(row=row_num, column=6, value='Weeks 0 to 2 - Lower CI')
    output_ws.cell(row=row_num, column=7, value='Weeks 0 to 2 - Upper CI')
    output_ws.cell(row=row_num, column=8, value='Weeks 2 to 4 - Mean')
    output_ws.cell(row=row_num, column=9, value='Weeks 2 to 4 - Lower CI')
    output_ws.cell(row=row_num, column=10, value='Weeks 2 to 4 - Upper CI')
    output_ws.cell(row=row_num, column=11, value='Weeks 4 to 12 - Mean')
    output_ws.cell(row=row_num, column=12, value='Weeks 4 to 12 - Lower CI')
    output_ws.cell(row=row_num, column=13, value='Weeks 4 to 12 - Upper CI')
    output_ws.cell(row=row_num, column=14, value='Weeks 12+ - Mean')
    output_ws.cell(row=row_num, column=15, value='Weeks 12+ - Lower CI')
    output_ws.cell(row=row_num, column=16, value='Weeks 12+ - Upper CI')
    row_num += 1

    output_ws.cell(row=row_num, column=1, value='Number of Shift Changes')
    output_ws.cell(row=row_num, column=2, value=shift_changes_stats_total.mean)
    output_ws.cell(row=row_num, column=3, value=shift_changes_stats_total.ci_lower)
    output_ws.cell(row=row_num, column=4, value=shift_changes_stats_total.ci_upper)
    output_ws.cell(row=row_num, column=5, value=shift_changes_stats_0.mean)
    output_ws.cell(row=row_num, column=6, value=shift_changes_stats_0.ci_lower)
    output_ws.cell(row=row_num, column=7, value=shift_changes_stats_0.ci_upper)
    output_ws.cell(row=row_num, column=8, value=shift_changes_stats_2.mean)
    output_ws.cell(row=row_num, column=9, value=shift_changes_stats_2.ci_lower)
    output_ws.cell(row=row_num, column=10, value=shift_changes_stats_2.ci_upper)
    output_ws.cell(row=row_num, column=11, value=shift_changes_stats_4.mean)
    output_ws.cell(row=row_num, column=12, value=shift_changes_stats_4.ci_lower)
    output_ws.cell(row=row_num, column=13, value=shift_changes_stats_4.ci_upper)
    output_ws.cell(row=row_num, column=14, value=shift_changes_stats_12.mean)
    output_ws.cell(row=row_num, column=15, value=shift_changes_stats_12.ci_lower)
    output_ws.cell(row=row_num, column=16, value=shift_changes_stats_12.ci_upper)
    
    # Create a new worksheet if necessary
    if len(output_wb.worksheets) < num_scenarios:
        output_ws = output_wb.create_sheet()

# Save the workbook results
current_loc = Path('.')
save_loc = (current_loc / 'results' / f'simulation_results_{int(time.time())}.xlsx').resolve()
print(f'Writing results to file: {save_loc}')
output_wb.save(save_loc)